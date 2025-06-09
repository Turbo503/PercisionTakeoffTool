import sys
from collections import defaultdict
from typing import Any
import os
import traceback
from pathlib import Path
from openpyxl import Workbook

from PyQt5 import QtCore, QtGui, QtWidgets
import fitz  # PyMuPDF – `pip install PyMuPDF`

# ─────────────────────────────────────────────────────────────────────────────
# Colour options (human‑readable → hex)
# ─────────────────────────────────────────────────────────────────────────────
color_options = {
    "Red": "#FF0000",      "Green": "#00FF00",     "Blue": "#0000FF",
    "Yellow": "#FFFF00",   "Magenta": "#FF00FF",   "Cyan": "#00FFFF",
    "Maroon": "#800000",   "Dark Green": "#008000","Navy": "#000080",
    "Olive": "#808000",    "Purple": "#800080",    "Teal": "#008080",
    "Silver": "#C0C0C0",   "Orange": "#FFA500",    "Brown": "#A52A2A",
    "Burly Wood": "#DEB887","Cadet Blue": "#5F9EA0","Chartreuse": "#7FFF00",
    "Chocolate": "#D2691E","Coral": "#FF7F50",    "Cornflower Blue": "#6495ED",
    "Crimson": "#DC143C",  "Dark Turquoise": "#00CED1","Dark Violet": "#9400D3",
    "Gold": "#FFD700",
}

# ─────────────────────────────────────────────────────────────────────────────
# HighlightItem – movable/deletable rectangle
# ─────────────────────────────────────────────────────────────────────────────
class HighlightItem(QtWidgets.QGraphicsObject):
    deleted = QtCore.pyqtSignal(object)

    def __init__(self, rect: QtCore.QRectF, color: QtGui.QColor, parent=None):
        super().__init__(parent)
        self._rect = rect
        self._color = color
        self.setFlags(QtWidgets.QGraphicsItem.ItemIsSelectable)

    # QGraphicsObject interface --------------------------------------------
    def boundingRect(self) -> QtCore.QRectF:
        return self._rect

    def paint(self, painter: QtGui.QPainter, _option, _widget):
        painter.setBrush(QtGui.QBrush(self._color))
        painter.setPen(QtCore.Qt.NoPen)
        painter.drawRect(self._rect)

    # helpers ---------------------------------------------------------------
    def setRect(self, rect: QtCore.QRectF):
        self.prepareGeometryChange()
        self._rect = rect
        self.update()

    # context menu ----------------------------------------------------------
    def contextMenuEvent(self, event: QtWidgets.QGraphicsSceneContextMenuEvent):
        menu = QtWidgets.QMenu()
        move_action = menu.addAction("Move")
        delete_action = menu.addAction("Delete")
        action = menu.exec_(event.screenPos())
        view: "PDFGraphicsView" = self.scene().views()[0]  # type: ignore
        if action == move_action:
            view.startMovingItem(self)
        elif action == delete_action:
            view.highlightDeleted.emit(self)
            self.scene().removeItem(self)
        event.accept()


# ─────────────────────────────────────────────────────────────────────────────
# PDFGraphicsView – page display, draw/stamp, panning, zoom, D‑n‑D
# ─────────────────────────────────────────────────────────────────────────────
class PDFGraphicsView(QtWidgets.QGraphicsView):
    stampDropped      = QtCore.pyqtSignal(object)
    highlightDeleted  = QtCore.pyqtSignal(object)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setDragMode(QtWidgets.QGraphicsView.ScrollHandDrag)
        self.setMouseTracking(True)
        self.setAcceptDrops(True)                             # NEW
        self._scene = QtWidgets.QGraphicsScene(self)
        self.setScene(self._scene)

        # drawing helpers
        self.draw_mode          = False
        self.drawing            = False
        self.template_defined   = False
        self.template_item: QtWidgets.QGraphicsRectItem | None = None
        self.start_point: QtCore.QPointF | None = None
        self.current_highlight_color = QtGui.QColor(255, 0, 0, 100)

        # move helpers
        self.moving_item: HighlightItem | None = None
        self.current_takeoff: dict | None = None

        # pdf
        self.doc: fitz.Document | None = None
        self.current_page: int = -1

    # ── drawing‑mode toggle ────────────────────────────────────────────
    def setDrawingMode(self, enabled: bool):
        self.draw_mode = enabled
        if enabled:
            self.setCursor(QtCore.Qt.CrossCursor)
            self.setDragMode(QtWidgets.QGraphicsView.NoDrag)
        else:
            self.setCursor(QtCore.Qt.ArrowCursor)
            self.setDragMode(QtWidgets.QGraphicsView.ScrollHandDrag)
            if self.template_item:
                self._scene.removeItem(self.template_item)
                self.template_item = None
            self.drawing = False
            self.template_defined = False

    # ------------------------------------------------------------------
    def cloneTemplate(self, item: QtWidgets.QGraphicsRectItem) -> HighlightItem:
        new_item = HighlightItem(item.rect(), self.current_highlight_color)
        new_item.deleted.connect(self.handleHighlightDeleted)
        new_item.page = self.current_page  # type: ignore[attr-defined]
        return new_item

    # ------------------------------------------------------------------
    def display_page(self, page_num: int):
        if not self.doc or page_num < 0 or page_num >= self.doc.page_count:
            return
        if hasattr(self, "_pixmap_item") and self._pixmap_item:
            self._scene.removeItem(self._pixmap_item)         # type: ignore[attr-defined]
        page = self.doc.load_page(page_num)
        pix  = page.get_pixmap()
        img  = QtGui.QImage(pix.samples, pix.width, pix.height,
                            pix.stride, QtGui.QImage.Format_RGB888)
        pixmap = QtGui.QPixmap.fromImage(img)
        self._pixmap_item = self._scene.addPixmap(pixmap)     # type: ignore[attr-defined]
        self._pixmap_item.setZValue(-10)                      # type: ignore[attr-defined]
        self.setSceneRect(QtCore.QRectF(pixmap.rect()))
        self.current_page = page_num

    # ------------------------------------------------------------------
    def load_pdf(self, pdf_path: str):
        try:
            self.doc = fitz.open(pdf_path)
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, "Error",
                                          f"Failed to open PDF:\n{e}")
            return
        self._scene.clear()
        if self.doc.page_count:
            self.display_page(0)

    # ------------------------------------------------------------------
    # zoom with wheel ---------------------------------------------------
    def wheelEvent(self, event: QtGui.QWheelEvent):
        z_in, z_out = 1.25, 1 / 1.25
        self.scale(z_in if event.angleDelta().y() > 0 else z_out,
                   z_in if event.angleDelta().y() > 0 else z_out)
        event.accept()

    # ------------------------------------------------------------------
    # fix ghost‑rectangle during keyboard pan --------------------------  NEW
    def scrollContentsBy(self, dx: int, dy: int):
        super().scrollContentsBy(dx, dy)
        if self.template_item:
            self.viewport().update()

    # ------------------------------------------------------------------
    # drag‑and‑drop support (PDF files) -------------------------------- NEW
    def dragEnterEvent(self, event: QtGui.QDragEnterEvent):
        if event.mimeData().hasUrls():
            for url in event.mimeData().urls():
                if url.toLocalFile().lower().endswith(".pdf"):
                    event.acceptProposedAction()
                    return
        super().dragEnterEvent(event)

    def dropEvent(self, event: QtGui.QDropEvent):
        urls = event.mimeData().urls()
        if not urls:
            return
        file_path = urls[0].toLocalFile()
        if not file_path.lower().endswith(".pdf"):
            return

        # let the main window know (if it exists) so thumbnails update
        mw = self.window()
        if hasattr(mw, "populateThumbnails"):
            mw.last_dir = QtCore.QFileInfo(file_path).absolutePath()
            mw.settings.setValue("lastDir", mw.last_dir)
            mw.pdf_file = file_path
            self.load_pdf(file_path)          # view first so it’s instant
            mw.populateThumbnails()           # type: ignore
        else:
            self.load_pdf(file_path)

        event.acceptProposedAction()

    # ------------------------------------------------------------------
    # ───── Mouse handling (draw / stamp / move) ───────────────────────
    def mousePressEvent(self, event: QtGui.QMouseEvent):
        # finish move ---------------------------------------------------
        if self.moving_item:
            if event.button() == QtCore.Qt.LeftButton:
                if self.current_takeoff is not None:
                    self.current_takeoff["highlights"].append(self.moving_item)
                    self.stampDropped.emit(self.moving_item)
                self.moving_item = None
                return
            elif event.button() == QtCore.Qt.RightButton:
                self._scene.removeItem(self.moving_item)
                self.highlightDeleted.emit(self.moving_item)
                self.moving_item = None
                return

        # cancel draw‑mode with right‑click (unless on highlight) -------
        if self.draw_mode and event.button() == QtCore.Qt.RightButton:
            if isinstance(self.itemAt(event.pos()), HighlightItem):
                super().mousePressEvent(event)
            else:
                self.setDrawingMode(False)
            return

        # start drawing -------------------------------------------------
        if self.draw_mode and event.button() == QtCore.Qt.LeftButton:
            if not self.template_defined:
                self.drawing = True
                self.start_point = self.mapToScene(event.pos())
                self.template_item = QtWidgets.QGraphicsRectItem(
                    QtCore.QRectF(self.start_point, self.start_point))
                self.template_item.setBrush(QtGui.QBrush(self.current_highlight_color))
                self.template_item.setPen(QtGui.QPen(QtCore.Qt.NoPen))
                self._scene.addItem(self.template_item)
                return
            else:
                stamped = self.cloneTemplate(self.template_item)        # type: ignore
                stamped.setRect(self.template_item.rect())              # type: ignore
                self._scene.addItem(stamped)
                self.stampDropped.emit(stamped)
                return

        super().mousePressEvent(event)

    # ------------------------------------------------------------------
    def mouseMoveEvent(self, event: QtGui.QMouseEvent):
        if self.moving_item:
            return super().mouseMoveEvent(event)  # handled in press/release

        # first rectangle definition
        if self.draw_mode and self.drawing and self.template_item:
            p = self.mapToScene(event.pos())
            self.template_item.setRect(QtCore.QRectF(self.start_point, p).normalized())
            return

        # template follows cursor
        if self.draw_mode and self.template_defined and not self.drawing and self.template_item:
            p = self.mapToScene(event.pos())
            size = self.template_item.rect().size()
            self.template_item.setRect(QtCore.QRectF(p, size))
            return

        super().mouseMoveEvent(event)

    # ------------------------------------------------------------------
    def mouseReleaseEvent(self, event: QtGui.QMouseEvent):
        if self.draw_mode and event.button() == QtCore.Qt.LeftButton and self.drawing:
            self.drawing = False
            self.template_defined = True
            stamped = self.cloneTemplate(self.template_item)             # type: ignore
            stamped.setRect(self.template_item.rect())                  # type: ignore
            self._scene.addItem(stamped)
            self.stampDropped.emit(stamped)
            return
        super().mouseReleaseEvent(event)

    # ------------------------------------------------------------------
    # begin move/stamp based on existing item
    def startMovingItem(self, item: HighlightItem):
        self.setDrawingMode(True)
        self.template_defined = True
        self.drawing = False

        if self.template_item and self.template_item is not item and self.template_item.scene():
            self._scene.removeItem(self.template_item)

        self.template_item = item
        self.current_highlight_color = item._color
        self.moving_item = None

    # ------------------------------------------------------------------
    def handleHighlightDeleted(self, item: HighlightItem):
        if self.current_takeoff and item in self.current_takeoff["highlights"]:
            self.current_takeoff["highlights"].remove(item)


# ─────────────────────────────────────────────────────────────────────────────
# TakeoffPanel – UI for per‑category take‑offs
# ─────────────────────────────────────────────────────────────────────────────
class TakeoffPanel(QtWidgets.QWidget):
    new_takeoff_signal = QtCore.pyqtSignal(dict)
    saveRequested      = QtCore.pyqtSignal()
    importRequested    = QtCore.pyqtSignal()
    totalsUpdated      = QtCore.pyqtSignal()

    def __init__(self, parent=None, *, include_wire: bool = True):
        super().__init__(parent)
        self.include_wire = include_wire

        main_layout = QtWidgets.QVBoxLayout(self)
        main_layout.setContentsMargins(2, 2, 2, 2)
        main_layout.setSpacing(4)

        # control row ----------------------------------------------------
        ctrl = QtWidgets.QHBoxLayout()
        self.add_btn   = QtWidgets.QPushButton("Add Takeoff")
        self.save_btn  = QtWidgets.QPushButton("Save")
        ctrl.addWidget(self.add_btn)
        ctrl.addWidget(self.save_btn)
        main_layout.addLayout(ctrl)
        self.add_btn.clicked.connect(self.add_takeoff)
        self.save_btn.clicked.connect(self.saveRequested.emit)

        # totals ---------------------------------------------------------
        self.totals_label = QtWidgets.QLabel("Totals: Count=0; Hours=0.00")
        main_layout.addWidget(self.totals_label)

        # scroll area ----------------------------------------------------
        self.takeoff_items: list[dict] = []
        self.scroll = QtWidgets.QScrollArea()
        self.scroll.setWidgetResizable(True)
        main_layout.addWidget(self.scroll)
        container = QtWidgets.QWidget()
        self.scroll.setWidget(container)
        self.container_layout = QtWidgets.QVBoxLayout(container)
        self.container_layout.setAlignment(QtCore.Qt.AlignTop)
        self.pdf_view: PDFGraphicsView | None = None

    # ------------------------------------------------------------------
    def setPdfView(self, view: "PDFGraphicsView"):
        self.pdf_view = view

    # ------------------------------------------------------------------
    def add_takeoff(self):
        frame = QtWidgets.QFrame()
        frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        layout = QtWidgets.QVBoxLayout(frame)

        # Row 1 – label/draw/colour ----------------------------------
        r1 = QtWidgets.QHBoxLayout()
        lbl      = QtWidgets.QLabel(f"Takeoff {len(self.takeoff_items) + 1}")
        draw_btn = QtWidgets.QPushButton("Draw Highlight")
        color_cb = QtWidgets.QComboBox()
        color_cb.addItems(color_options.keys())
        r1.addWidget(lbl); r1.addWidget(draw_btn); r1.addWidget(color_cb)
        layout.addLayout(r1)

        # Row 2 – details --------------------------------------------
        r2 = QtWidgets.QHBoxLayout()
        count_lbl  = QtWidgets.QLabel("Count: 0")
        name_f     = QtWidgets.QLineEdit(); name_f.setPlaceholderText("Name")
        labor_lbl  = QtWidgets.QLabel("Labor:")
        labor_f    = QtWidgets.QLineEdit(); labor_f.setPlaceholderText("0.0")
        labor_f.setFixedWidth(60)
        labor_f.textChanged.connect(self.update_totals)
        r2.addWidget(count_lbl); r2.addWidget(name_f)
        r2.addWidget(labor_lbl); r2.addWidget(labor_f)

        if self.include_wire:
            type_cb   = QtWidgets.QComboBox(); type_cb.addItems(["NMD", "AC90", "SOW", "SJOW"])
            cable_cb  = QtWidgets.QComboBox(); cable_cb.addItems([
                "14-2","14-3","14-4","12-2","12-3","12-4","10-2","10-3","10-4",
                "8-2","8-3","8-4","6-2","6-3","6-4","4-2","4-3","4-4",
                "2-2","2-3","2-4","1/0-3","1/0-4","2/0-3","2/0-4","3/0-3","3/0-4",
                "4/0-3","4/0-4","250 MCM","300 MCM","350 MCM","500 MCM",
            ])
            mat_cb    = QtWidgets.QComboBox(); mat_cb.addItems(["CU", "AL"])
            length_lbl= QtWidgets.QLabel("Length:")
            length_f  = QtWidgets.QLineEdit(); length_f.setPlaceholderText("0.0")
            length_f.setFixedWidth(60)
            length_f.textChanged.connect(self.update_totals)
            r2.addWidget(type_cb); r2.addWidget(cable_cb); r2.addWidget(mat_cb)
            r2.addWidget(length_lbl); r2.addWidget(length_f)

        del_btn   = QtWidgets.QPushButton("Delete")
        notes_btn = QtWidgets.QPushButton("Show Notes")
        r2.addWidget(del_btn); r2.addWidget(notes_btn)
        layout.addLayout(r2)

        # Row 3 – notes ----------------------------------------------
        note_w = QtWidgets.QPlainTextEdit(); note_w.setFixedHeight(100)
        note_w.setVisible(False)
        layout.addWidget(note_w)

        self.container_layout.addWidget(frame)

        # store references -------------------------------------------
        item: dict[str, Any] = {
            "frame": frame, "label": lbl, "draw": draw_btn, "color": color_cb,
            "count": count_lbl, "name": name_f, "labor": labor_f,
            "notes": note_w, "highlights": [],
            # aliases for legacy code
            "color_dropdown": color_cb, "name_field": name_f,
            "labor_field": labor_f, "notes_edit": note_w,
        }
        if self.include_wire:
            item.update({
                "wire_type": type_cb, "wire_cable": cable_cb,
                "wire_mat": mat_cb, "wire_length": length_f,
            })
        self.takeoff_items.append(item)

        # connections -------------------------------------------------
        draw_btn.clicked.connect(lambda _=False, it=item: self.new_takeoff_signal.emit(it))
        del_btn.clicked.connect(lambda _=False, it=item: self.delete_takeoff(it))
        notes_btn.clicked.connect(lambda _=False, b=notes_btn, n=note_w: (
            n.setVisible(not n.isVisible()),
            b.setText("Hide Notes" if n.isVisible() else "Show Notes"),
        ))

        self.update_totals()

    # ------------------------------------------------------------------
    def delete_takeoff(self, item: dict):
        for h in item["highlights"]:
            if h.scene():
                h.scene().removeItem(h)
        item["frame"].setParent(None)
        self.takeoff_items.remove(item)
        self.update_totals()

    # ------------------------------------------------------------------
    def update_count(self, takeoff_item: dict):
        valid = [h for h in takeoff_item["highlights"] if h.scene()]
        takeoff_item["count"].setText(f"Count: {len(valid)}")
        self.update_totals()

    # ------------------------------------------------------------------
    def update_totals(self):
        total_count = 0
        total_hours = 0.0
        for it in self.takeoff_items:
            valid = [h for h in it["highlights"] if h.scene()]
            cnt   = len(valid)
            it["count"].setText(f"Count: {cnt}")
            total_count += cnt
            try:
                lab_widget = it.get("labor_field", it.get("labor"))
                lab = float(lab_widget.text()) if lab_widget else 0.0
            except Exception:
                lab = 0.0
            total_hours += cnt * lab
        self.totals_label.setText(f"Totals: Count={total_count}; Hours={total_hours:.2f}")
        self.totalsUpdated.emit()

    # ------------------------------------------------------------------
    def clearTakeoffs(self):
        while self.takeoff_items:
            self.delete_takeoff(self.takeoff_items[0])

    # ------------------------------------------------------------------
    # NEW – multiply wire length by device count
    def get_wire_totals(self) -> defaultdict[tuple, float]:
        totals: defaultdict[tuple, float] = defaultdict(float)
        if not self.include_wire:
            return totals

        for it in self.takeoff_items:
            try:
                base_len = float(it["wire_length"].text())
            except Exception:
                continue
            if base_len <= 0:
                continue
            count = len([h for h in it["highlights"] if h.scene()])
            if count == 0:
                continue
            key = (
                it["wire_type"].currentText(),
                it["wire_cable"].currentText(),
                it["wire_mat"].currentText(),
            )
            totals[key] += base_len * count
        return totals


# ─────────────────────────────────────────────────────────────────────────────
# MainWindow – thumbnails / view / take‑offs / summary
# ─────────────────────────────────────────────────────────────────────────────
class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("PDF Takeoff Tool")
        self.resize(1200, 800)

        # persistent settings (must be first for D‑n‑D)
        self.settings = QtCore.QSettings("YourCompany", "PDFTakeoffTool")
        self.last_dir = self.settings.value("lastDir", QtCore.QDir.homePath())

        splitter = QtWidgets.QSplitter(QtCore.Qt.Horizontal)
        self.setCentralWidget(splitter)

        # left – thumbnails ---------------------------------------------
        self.thumbnail_list = QtWidgets.QListWidget()
        self.thumbnail_list.setViewMode(QtWidgets.QListView.IconMode)
        self.thumbnail_list.setIconSize(QtCore.QSize(200, 200))
        self.thumbnail_list.setResizeMode(QtWidgets.QListView.Adjust)
        self.thumbnail_list.setSpacing(5)
        splitter.addWidget(self.thumbnail_list)

        # centre – PDF view --------------------------------------------
        self.pdf_view = PDFGraphicsView()
        splitter.addWidget(self.pdf_view)

        # right – summary + tabs ---------------------------------------
        right = QtWidgets.QWidget()
        rlay  = QtWidgets.QVBoxLayout(right); rlay.setContentsMargins(2, 2, 2, 2); rlay.setSpacing(4)

        sf = QtWidgets.QFrame()
        sl = QtWidgets.QHBoxLayout(sf)
        self.sum_hours   = QtWidgets.QLabel("Total Hours: 0.00")
        self.sum_devices = QtWidgets.QLabel("Total Devices: 0")
        self.sum_wire    = QtWidgets.QLabel("Wire Totals: -")
        sl.addWidget(self.sum_hours); sl.addWidget(self.sum_devices); sl.addWidget(self.sum_wire)
        rlay.addWidget(sf)

        self.category_tabs = QtWidgets.QTabWidget()
        rlay.addWidget(self.category_tabs, 1)
        splitter.addWidget(right)

        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 7)
        splitter.setStretchFactor(2, 2)
        splitter.setSizes([120, 900, 180])

        # panels --------------------------------------------------------
        self.panels: dict[str, TakeoffPanel] = {}
        for name, use_wire in [("General", True), ("Lighting", True),
                               ("Mechanical", True), ("Demo", False)]:
            p = TakeoffPanel(include_wire=use_wire)
            p.setPdfView(self.pdf_view)
            self.category_tabs.addTab(p, name)
            self.panels[name] = p
            p.new_takeoff_signal.connect(self.start_draw_for_takeoff)
            p.saveRequested.connect(self.save)
            p.totalsUpdated.connect(self.update_summary)

        # signals -------------------------------------------------------
        self.pdf_view.stampDropped.connect(self.handleStampDropped)
        self.pdf_view.highlightDeleted.connect(self.handleHighlightDeleted)
        self.thumbnail_list.itemClicked.connect(self.thumbnailClicked)
        self.setAcceptDrops(True)

        # menu ----------------------------------------------------------
        file_menu = self.menuBar().addMenu("File")
        open_act  = QtWidgets.QAction("Open PDF", self); file_menu.addAction(open_act)
        open_act.triggered.connect(self.open_pdf_dialog)

        self.current_takeoff: dict | None = None

    # ------------------------------------------------------------------
    def start_draw_for_takeoff(self, takeoff_item: dict):
        self.current_takeoff = takeoff_item
        self.pdf_view.current_takeoff = takeoff_item

        color_name = takeoff_item["color"].currentText()
        col = QtGui.QColor(color_options[color_name]); col.setAlpha(80)
        self.pdf_view.current_highlight_color = col
        self.pdf_view.setDrawingMode(True)

    # ------------------------------------------------------------------
    def update_summary(self):
        total_hours, total_devices = 0.0, 0
        wire_totals: defaultdict[tuple, float] = defaultdict(float)

        for name, panel in self.panels.items():
            for it in panel.takeoff_items:
                cnt = len([h for h in it["highlights"] if h.scene()])
                try:
                    lab = float(it["labor_field"].text())
                except Exception:
                    lab = 0.0
                total_hours += cnt * lab
                if name != "Demo":
                    total_devices += cnt
            if panel.include_wire and name != "Demo":
                for key, length in panel.get_wire_totals().items():
                    wire_totals[key] += length

        self.sum_hours.setText(f"Total Hours: {total_hours:.2f}")
        self.sum_devices.setText(f"Total Devices: {total_devices}")
        self.sum_wire.setText(
            "Wire Totals: " + (
                "; ".join(f"{length:.2f} {t}/{c} ({m})"
                          for (t, c, m), length in wire_totals.items())
                if wire_totals else "-"
            )
        )

    # ------------------------------------------------------------------
    def resizeEvent(self, event: QtGui.QResizeEvent):
        super().resizeEvent(event)
        width  = self.thumbnail_list.viewport().width() - 4
        height = int(width * 1.414)   # ISO A‑series ratio
        self.thumbnail_list.setIconSize(QtCore.QSize(width, height))

    # ------------------------------------------------------------------
    # drag‑and‑drop onto whole window (legacy)
    def dragEnterEvent(self, event: QtGui.QDragEnterEvent):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEvent(self, event: QtGui.QDropEvent):
        urls = event.mimeData().urls()
        if not urls:
            return
        file_path = urls[0].toLocalFile()
        if not file_path.lower().endswith(".pdf"):
            return
        self.last_dir = QtCore.QFileInfo(file_path).absolutePath()
        self.settings.setValue("lastDir", self.last_dir)
        self.pdf_file = file_path
        self.pdf_view.load_pdf(file_path)
        self.populateThumbnails()

    # ------------------------------------------------------------------
    def open_pdf_dialog(self):
        p, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, "Open PDF", self.last_dir, "PDF Files (*.pdf)")
        if not p:
            return
        self.last_dir = QtCore.QFileInfo(p).absolutePath()
        self.settings.setValue("lastDir", self.last_dir)
        self.pdf_file = p
        self.pdf_view.load_pdf(p)
        self.populateThumbnails()

    # ------------------------------------------------------------------
    def populateThumbnails(self):
        self.thumbnail_list.clear()
        if self.pdf_view.doc is None:
            return
        for i in range(self.pdf_view.doc.page_count):
            page = self.pdf_view.doc.load_page(i)
            pix  = page.get_pixmap(matrix=fitz.Matrix(0.2, 0.2))
            img  = QtGui.QImage(pix.samples, pix.width, pix.height,
                                pix.stride, QtGui.QImage.Format_RGB888)
            pixmap = QtGui.QPixmap.fromImage(img)
            icon   = QtGui.QIcon(pixmap)
            item   = QtWidgets.QListWidgetItem(icon, "")
            item.setData(QtCore.Qt.UserRole, i)
            self.thumbnail_list.addItem(item)

    # ------------------------------------------------------------------
    def thumbnailClicked(self, item: QtWidgets.QListWidgetItem):
        pg = item.data(QtCore.Qt.UserRole)
        self.pdf_view.display_page(pg)
        self.updateHighlightsForPage(pg)

    # ------------------------------------------------------------------
    def handleStampDropped(self, stamped):
        if self.current_takeoff is None:
            return
        self.current_takeoff["highlights"].append(stamped)
        for panel in self.panels.values():
            if self.current_takeoff in panel.takeoff_items:
                panel.update_count(self.current_takeoff)
                break

    # ------------------------------------------------------------------
    def handleHighlightDeleted(self, item):
        for panel in self.panels.values():
            for takeoff in panel.takeoff_items:
                if item in takeoff["highlights"]:
                    takeoff["highlights"].remove(item)
                    panel.update_count(takeoff)
                    return

    def updateHighlightsForPage(self, page_num: int):
        for panel in self.panels.values():
            for takeoff in panel.takeoff_items:
                for h in takeoff["highlights"]:
                    try:
                        visible = hasattr(h, "page") and int(h.page) == page_num
                    except Exception:
                        visible = False
                    try:
                        h.setVisible(visible)
                    except RuntimeError:
                        # catch objects already deleted from scene
                        continue

    # ------------------------------------------------------------------
    # SAVE – export to .ods (Summary, Highlights + per‑category sheets)
    # ------------------------------------------------------------------
    def save(self):  # (unchanged line)
        """
        Create a one‑sheet spreadsheet containing only the take‑off names that the
        user typed in the “Name” field for each highlight, and the count per takeoff.

        • All entries start in *column A, row 4*.
        • A completely blank row separates each tab’s section.
        • A single row with the literal text “Labor” in column A and the *total labor* in column B.
        • The file is saved next to the PDF that was dragged‑and‑dropped, using the
          same base‑name and the extension “.xlsx”.
        """
        import traceback

        try:
            # ------------------------------------------------------------------
            # 1.  Collect the take‑off names and counts per tab
            # ------------------------------------------------------------------
            sections = []  # list[list[tuple[str, int]]]
            total_labor = 0.0

            for tab_index in range(self.category_tabs.count()):
                panel = self.category_tabs.widget(tab_index)
                entries = []
                for takeoff in panel.takeoff_items:
                    name_widget = takeoff.get("name_field")
                    name = name_widget.text().strip() if name_widget else ""
                    if not name:
                        continue
                    count = len([h for h in takeoff["highlights"] if h.scene()])
                    entries.append((name, count))
                    try:
                        labor = float(takeoff.get("labor_field").text())
                    except Exception:
                        labor = 0.0
                    total_labor += labor * count
                sections.append(entries)

            # ------------------------------------------------------------------
            # 2.  Build the workbook entirely in memory
            # ------------------------------------------------------------------
            wb = Workbook()
            ws = wb.active
            ws.title = "Estimate"

            row = 4  # first row for data
            for idx, section in enumerate(sections, start=1):
                for name, count in section:
                    ws.cell(row=row, column=1, value=name)
                    ws.cell(row=row, column=2, value=count)
                    row += 1

                # Blank divider row between tabs (except after last one)
                if idx < len(sections):
                    row += 1

            # Final "Labor" total row
            ws.cell(row=row, column=1, value="Labor")
            ws.cell(row=row, column=2, value=round(total_labor, 2))

            # ------------------------------------------------------------------
            # 3.  Determine the destination file name
            # ------------------------------------------------------------------
            pdf_path = getattr(self, "pdf_file", None)
            if not pdf_path:
                QtWidgets.QMessageBox.warning(
                    self, "Save failed", "No PDF is currently loaded."
                )
                return

            out_path = Path(pdf_path).with_suffix(".xlsx")

            # ------------------------------------------------------------------
            # 4.  Write to disk and confirm
            # ------------------------------------------------------------------
            wb.save(out_path)
            QtWidgets.QMessageBox.information(
                self,
                "Saved",
                f"Spreadsheet written to:\n{out_path}",
            )

        except Exception as exc:
            msg = "".join(traceback.format_exception(exc))
            QtWidgets.QMessageBox.critical(self, "Save error", msg)
            print(msg, file=sys.stderr)


# ─────────────────────────────────────────────────────────────────────────────
# main entry‑point
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())