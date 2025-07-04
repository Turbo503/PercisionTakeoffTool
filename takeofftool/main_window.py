from __future__ import annotations
from pathlib import Path
from PyQt5 import QtCore, QtGui, QtWidgets
from .viewer import PDFGraphicsView, HighlightItem, LineItem
from .panels import TakeoffPanel, color_options
from PyQt5 import QtCore
import fitz, os, shutil, gc     # gc so we can be SURE nothing is hanging on
import json, base64, subprocess, tempfile, sys, os, textwrap, shutil

def launch_save_worker(dest_path: str, pdf_bytes: bytes, hl_dump: list):
    """Spawn save_worker.py in a clean Python process and wait for it."""
    # bundle everything into a temp json file so we avoid crazy cmd-line quoting
    bundle_path = tempfile.mktemp(suffix=".json")
    with open(bundle_path, "w", encoding="utf-8") as fp:
        fp.write(json.dumps({
            "pdf_hex": pdf_bytes.hex(),
            "hl": hl_dump,
        }))
    # `sys.executable` preserves the same venv / pip installs
    cmd = [sys.executable, os.path.join(os.path.dirname(__file__), "save_worker.py"),
           dest_path, bundle_path]
    proc = subprocess.run(cmd)
    os.remove(bundle_path)
    return proc.returncode == 0

def _gather_highlights(panels) -> list:
    """Return plain-Python highlight descriptors without importing fitz."""
    dump = []
    for panel in panels.values():
        for take in panel.takeoff_items:
            for h in take["highlights"]:
                page = getattr(h, "page", -1)
                if page < 0:
                    continue
                if isinstance(h, HighlightItem):
                    r = h.rect()
                    dump.append({
                        "kind": "rect",
                        "page": int(page),
                        "rect": [r.x(), r.y(), r.x() + r.width(), r.y() + r.height()],
                        "color": list(h._color.getRgbF()[:3]),
                    })
                elif isinstance(h, LineItem):
                    line = h.line()
                    dump.append({
                        "kind": "line",
                        "page": int(page),
                        "p1": [line.x1(), line.y1()],
                        "p2": [line.x2(), line.y2()],
                        "width": h.pen().widthF(),
                        "color": list(h.pen().color().getRgbF()[:3]),
                    })
    return dump


class RenderGuard(QtCore.QObject):
    """Context-manager that cleanly pauses your QThread renderer."""

    def __init__(self, wnd):
        super().__init__()
        self._wnd = wnd          # main window – must expose ._render_thread

    def __enter__(self):
        th = getattr(self._wnd, "_render_thread", None)
        if th and th.isRunning():
            th.requestInterruption()
            th.quit()
            th.wait()            # ← BLOCK until completely dead
        return self

    def __exit__(self, exc_type, exc, tb):
        # resurrect the thread if the window normally keeps it alive
        th = getattr(self._wnd, "_render_thread", None)
        if th:
            th.start()


class MainWindow(QtWidgets.QMainWindow):
    """Main application window."""

    def __init__(self):
        super().__init__()
        self.setWindowTitle("PDF Takeoff Tool")
        self.resize(1200, 800)

        self.settings = QtCore.QSettings("YourCompany", "PDFTakeoffTool")
        self.last_dir = self.settings.value("lastDir", QtCore.QDir.homePath())

        splitter = QtWidgets.QSplitter(QtCore.Qt.Horizontal)
        self.setCentralWidget(splitter)

        self.thumbnail_list = QtWidgets.QListWidget()
        self.thumbnail_list.setViewMode(QtWidgets.QListView.IconMode)
        self.thumbnail_list.setIconSize(QtCore.QSize(200, 200))
        self.thumbnail_list.setResizeMode(QtWidgets.QListView.Adjust)
        self.thumbnail_list.setSpacing(5)
        splitter.addWidget(self.thumbnail_list)

        self.pdf_view = PDFGraphicsView()
        splitter.addWidget(self.pdf_view)

        right = QtWidgets.QWidget()
        rlay = QtWidgets.QVBoxLayout(right)
        rlay.setContentsMargins(2, 2, 2, 2)
        rlay.setSpacing(4)
        sf = QtWidgets.QFrame()
        sl = QtWidgets.QHBoxLayout(sf)
        self.sum_hours = QtWidgets.QLabel("Total Hours: 0.00")
        self.sum_devices = QtWidgets.QLabel("Total Devices: 0")
        self.sum_points = QtWidgets.QLabel("Total Points: 0")
        sl.addWidget(self.sum_hours)
        sl.addWidget(self.sum_devices)
        sl.addWidget(self.sum_points)
        rlay.addWidget(sf)

        self.category_tabs = QtWidgets.QTabWidget()
        rlay.addWidget(self.category_tabs, 1)
        splitter.addWidget(right)

        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 7)
        splitter.setStretchFactor(2, 2)
        splitter.setSizes([120, 900, 180])

        self.panels: dict[str, TakeoffPanel] = {}
        for name in [
            "General",
            "Lighting",
            "Mechanical",
            "Fire Alarm",
            "Low Voltage",
            "Demo",
        ]:
            p = TakeoffPanel(include_wire=False)
            p.setPdfView(self.pdf_view)
            self.category_tabs.addTab(p, name)
            self.panels[name] = p
            p.new_takeoff_signal.connect(self.start_draw_for_takeoff)
            p.saveRequested.connect(self.save_excel)
            p.totalsUpdated.connect(self.update_summary)

        self.pdf_view.stampDropped.connect(self.handleStampDropped)
        self.pdf_view.highlightDeleted.connect(self.handleHighlightDeleted)
        self.thumbnail_list.itemClicked.connect(self.thumbnailClicked)
        self.setAcceptDrops(True)

        file_menu = self.menuBar().addMenu("File")
        open_act = QtWidgets.QAction("Open PDF", self)
        save_act = QtWidgets.QAction("Save PDF", self)
        save_as_act = QtWidgets.QAction("Save PDF As", self)
        file_menu.addAction(open_act)
        file_menu.addAction(save_act)
        file_menu.addAction(save_as_act)
        open_act.triggered.connect(self.open_pdf_dialog)
        save_act.triggered.connect(self.save_pdf)
        save_as_act.triggered.connect(self.save_pdf_as)

        self.pdf_file: str | None = None
        self.current_takeoff: dict | None = None
        self._original_pdf_bytes: bytes | None = None

    def start_draw_for_takeoff(self, takeoff_item: dict):
        self.current_takeoff = takeoff_item
        self.pdf_view.current_takeoff = takeoff_item
        color_name = takeoff_item["color"].currentText()
        col = QtGui.QColor(color_options[color_name])
        col.setAlpha(80)
        self.pdf_view.current_highlight_color = col
        self.pdf_view.setDrawingShape("rect")
        self.pdf_view.setDrawingMode(True)

    def update_summary(self):
        total_hours, total_devices, total_points = 0.0, 0, 0
        for name, panel in self.panels.items():
            for it in panel.takeoff_items:
                cnt = len([h for h in it["highlights"] if h.scene()])
                try:
                    lab = float(it["labor_field"].text())
                except Exception:
                    lab = 0.0
                total_hours += cnt * lab
                if name != "Demo":
                    total_devices += cnt
                total_points += cnt
        self.sum_hours.setText(f"Total Hours: {total_hours:.2f}")
        self.sum_devices.setText(f"Total Devices: {total_devices}")
        self.sum_points.setText(f"Total Points: {total_points}")

    # Thumbnail handling -------------------------------------------------
    def populateThumbnails(self):
        self.thumbnail_list.clear()
        if self.pdf_view.doc is None:
            return
        for i in range(self.pdf_view.doc.page_count):
            page = self.pdf_view.doc.load_page(i)
            pix = page.get_pixmap(matrix=fitz.Matrix(0.2, 0.2))
            img = QtGui.QImage(
                pix.samples, pix.width, pix.height, pix.stride, QtGui.QImage.Format_RGB888
            )
            pixmap = QtGui.QPixmap.fromImage(img)
            icon = QtGui.QIcon(pixmap)
            item = QtWidgets.QListWidgetItem(icon, "")
            item.setData(QtCore.Qt.UserRole, i)
            self.thumbnail_list.addItem(item)

    def thumbnailClicked(self, item: QtWidgets.QListWidgetItem):
        pg = item.data(QtCore.Qt.UserRole)
        self.pdf_view.display_page(pg)
        self.updateHighlightsForPage(pg)

    # Drag and drop -----------------------------------------------------
    def dragEnterEvent(self, event: QtGui.QDragEnterEvent):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEvent(self, event: QtGui.QDropEvent):
        urls = event.mimeData().urls()
        if not urls:
            return
        file_path = urls[0].toLocalFile()
        if not file_path.lower().endswith(".pdf"):
            return
        self.last_dir = QtCore.QFileInfo(file_path).absolutePath()
        self.settings.setValue("lastDir", self.last_dir)
        # ── inside dropEvent() *right after* you set self.pdf_file ─────────
        self.pdf_file = file_path
        with open(file_path, "rb") as fp:
            self._original_pdf_bytes = fp.read()  # ← NEW
        self.pdf_view.load_pdf(file_path)
        self.populateThumbnails()

    def open_pdf_dialog(self):
        p, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, "Open PDF", self.last_dir, "PDF Files (*.pdf)"
        )
        if not p:
            return
        self.last_dir = QtCore.QFileInfo(p).absolutePath()
        self.settings.setValue("lastDir", self.last_dir)
        # ── inside open_pdf_dialog() *right after* you get 'p' ─────────────
        self.pdf_file = p
        with open(p, "rb") as fp:
            self._original_pdf_bytes = fp.read()  # ← NEW
        self.pdf_view.load_pdf(p)
        self.populateThumbnails()

    # Stamps ------------------------------------------------------------
    def handleStampDropped(self, stamped):
        if self.current_takeoff is None:
            return
        self.current_takeoff["highlights"].append(stamped)
        for panel in self.panels.values():
            if self.current_takeoff in panel.takeoff_items:
                panel.update_count(self.current_takeoff)
                break

    def handleHighlightDeleted(self, item):
        for panel in self.panels.values():
            for takeoff in panel.takeoff_items:
                if item in takeoff["highlights"]:
                    takeoff["highlights"].remove(item)
                    panel.update_count(takeoff)
                    return

    def updateHighlightsForPage(self, page_num: int):
        for panel in self.panels.values():
            for takeoff in panel.takeoff_items:
                for h in takeoff["highlights"]:
                    try:
                        visible = hasattr(h, "page") and int(h.page) == page_num
                    except Exception:
                        visible = False
                    try:
                        h.setVisible(visible)
                    except RuntimeError:
                        continue

    # Excel export ------------------------------------------------------
    def save_excel(self):
        from openpyxl import Workbook

        try:
            sections = []
            total_labor = 0.0
            for tab_index in range(self.category_tabs.count()):
                panel = self.category_tabs.widget(tab_index)
                entries = []
                for takeoff in panel.takeoff_items:
                    name_widget = takeoff.get("name_field")
                    name = name_widget.text().strip() if name_widget else ""
                    if not name:
                        continue
                    count = len([h for h in takeoff["highlights"] if h.scene()])
                    entries.append((name, count))
                    try:
                        labor = float(takeoff.get("labor_field").text())
                    except Exception:
                        labor = 0.0
                    total_labor += labor * count
                sections.append(entries)

            wb = Workbook()
            ws = wb.active
            ws.title = "Estimate"

            row = 4
            for idx, section in enumerate(sections, start=1):
                for name, count in section:
                    ws.cell(row=row, column=1, value=name)
                    ws.cell(row=row, column=2, value=count)
                    row += 1
                if idx < len(sections):
                    row += 1
            ws.cell(row=row, column=1, value="Labor")
            ws.cell(row=row, column=2, value=round(total_labor, 2))

            pdf_path = self.pdf_file
            if not pdf_path:
                QtWidgets.QMessageBox.warning(self, "Save failed", "No PDF loaded")
                return
            out_path = Path(pdf_path).with_suffix(".xlsx")
            wb.save(out_path)
            QtWidgets.QMessageBox.information(self, "Saved", f"Spreadsheet written to:\n{out_path}")
        except Exception as exc:  # pragma: no cover - GUI feedback
            QtWidgets.QMessageBox.critical(self, "Save error", str(exc))

    def save_pdf(self):  # header stays
        """
        Overwrite the currently-open PDF with in-memory highlights.
        Zero MuPDF calls in this process.
        """
        from PyQt5 import QtWidgets
        import gc

        if not self.pdf_view.doc or not self.pdf_file:
            QtWidgets.QMessageBox.warning(self, "Save failed", "No PDF loaded.")
            return

        # 1️⃣  pristine bytes cached at load time
        pdf_bytes = self._original_pdf_bytes
        if pdf_bytes is None:  # extreme fallback
            with open(self.pdf_file, "rb") as fp:
                pdf_bytes = fp.read()

        # 2️⃣  snapshot highlights while render thread is *paused*
        with RenderGuard(self):
            hl_dump = _gather_highlights(self.panels)

        ok = launch_save_worker(self.pdf_file, pdf_bytes, hl_dump)

        if ok:
            QtWidgets.QMessageBox.information(self, "Saved",
                                              f"PDF overwritten:\n{self.pdf_file}")
        else:
            QtWidgets.QMessageBox.critical(
                self, "Error",
                "The external save process crashed. "
                "Your original document is unchanged."
            )
        gc.collect()

    # ───────────────────────────────────────────────────────────────
    def save_pdf_as(self):  # ←--- header kept, body replaced
        """
        "Save As…" convenience wrapper – delegates heavy-lifting to _export_pdf.
        """
        from PyQt5 import QtWidgets
        p, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "Save PDF As", self.last_dir, "PDF files (*.pdf)"
        )
        if p:
            self._export_pdf(Path(p))

    # ───────────────────────────────────────────────────────────────

    def _export_pdf(self, dest: Path):  # header unchanged
        """
        Save-As via helper process; GUI never links against MuPDF.
        """
        from PyQt5 import QtWidgets
        import gc

        if not self.pdf_view.doc:
            return

        # 1️⃣  bytes
        pdf_bytes = self._original_pdf_bytes
        if pdf_bytes is None and self.pdf_file:
            with open(self.pdf_file, "rb") as fp:
                pdf_bytes = fp.read()

        # 2️⃣  snapshot under RenderGuard
        with RenderGuard(self):
            hl_dump = _gather_highlights(self.panels)

        ok = launch_save_worker(str(dest), pdf_bytes, hl_dump)

        if ok:
            QtWidgets.QMessageBox.information(self, "Saved",
                                              f"PDF written to:\n{dest}")
        else:
            QtWidgets.QMessageBox.critical(
                self, "Error",
                "The external save process crashed. "
                "Your original document is unchanged."
            )
        gc.collect()

